"""
orders/export_views.py
======================
Export orders as a formatted Excel (.xlsx) file.
Admin only.

Add to orders/urls.py:
    from .export_views import export_orders_excel
    path('export/excel/', export_orders_excel, name='export-orders-excel'),

Requirements:
    pip install openpyxl

Query params:
    ?status=pending     filter by status
    ?days=7             last N days only
    ?format=items       include one row per item (default: one row per order)
"""

import io
from datetime import datetime, timedelta

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_orders_excel(request):
    if not request.user.is_staff:
        return Response({'error': 'Admin access required'}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'error': 'Run: pip install openpyxl'}, status=500)

    try:
        from orders.models import Order, OrderItem
    except ImportError:
        # Fallback model name guesses
        try:
            from orders.models import Order
            OrderItem = None
        except ImportError:
            return Response({'error': 'Order model not found'}, status=500)

    # ── Queryset ──
    qs = Order.objects.all().select_related('user')
    try:
        qs = qs.prefetch_related('items__product')
    except Exception:
        pass

    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    days = request.GET.get('days')
    if days:
        try:
            since = datetime.now() - timedelta(days=int(days))
            qs = qs.filter(created_at__gte=since)
        except ValueError:
            pass

    qs = qs.order_by('-created_at')

    # ── Workbook ──
    wb = openpyxl.Workbook()

    # ── Colors ──
    C_ORANGE = "FF6B35"
    C_GOLD   = "FFBE0B"
    C_DARK   = "1A1A2E"
    C_MID    = "16213E"
    C_LIGHT  = "F5F5F5"
    C_GREEN  = "27AE60"
    C_RED    = "E74C3C"
    C_YELLOW = "F39C12"
    C_GREY   = "BDC3C7"

    thin = Side(style='thin', color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(size=11, bold=True, color="FFFFFF"):
        return Font(name='Calibri', size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="1A1A1A"):
        return Font(name='Calibri', size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    def status_color(s):
        s = (s or '').lower()
        if s in ('completed', 'delivered', 'paid'):   return C_GREEN
        if s in ('cancelled', 'failed', 'rejected'):  return C_RED
        if s in ('pending',):                          return C_YELLOW
        return "5D6D7E"

    # ══════════════════════════════════════════
    # SHEET 1 — Orders Summary
    # ══════════════════════════════════════════
    ws = wb.active
    ws.title = "Orders"

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value     = f"📦  Pixels to Plastic — Orders Export  |  {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    title_cell.font      = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_cell.fill      = fill(C_ORANGE)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 36

    # Header row
    headers = [
        ("Order ID",   12), ("Date",          14), ("Customer",   18),
        ("Email",      22), ("Phone",          14), ("Items",       7),
        ("Total (₹)", 12), ("Payment",        12), ("Status",     12),
    ]
    for col, (hdr, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.font      = hdr_font()
        c.fill      = fill(C_DARK)
        c.alignment = center()
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, order in enumerate(qs, 3):
        user    = order.user
        name    = f"{getattr(user,'first_name','')} {getattr(user,'last_name','')}".strip() or user.username
        phone   = ''
        try:    phone = user.profile.phone
        except: pass

        # count items
        item_count = 0
        try:    item_count = order.items.count()
        except: pass

        try:    total = f"{order.total_price:,.2f}"
        except:
            try:    total = f"{order.total:,.2f}"
            except: total = '—'

        try:    payment = order.payment_method
        except: payment = '—'

        try:    order_status = order.status
        except: order_status = '—'

        try:    created = order.created_at.strftime('%d %b %Y  %H:%M')
        except: created = '—'

        row_data = [
            f"#{order.id}", created, name, user.email,
            phone, item_count, total, payment, order_status.capitalize()
        ]

        bg = C_LIGHT if row_idx % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font      = cell_font()
            c.fill      = fill(bg)
            c.border    = border
            c.alignment = center() if col in (1,5,6,7,8,9) else left()

        # Colour the status cell
        status_cell = ws.cell(row=row_idx, column=9)
        status_cell.fill = fill(status_color(order_status))
        status_cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")

        ws.row_dimensions[row_idx].height = 20

    # Totals row
    last = qs.count() + 3
    ws.merge_cells(f'A{last}:F{last}')
    ws[f'A{last}'].value      = f"Total Orders: {qs.count()}"
    ws[f'A{last}'].font       = hdr_font(size=10)
    ws[f'A{last}'].fill       = fill(C_DARK)
    ws[f'A{last}'].alignment  = left()

    # Freeze top 2 rows
    ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════
    # SHEET 2 — Order Items Detail
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet("Item Details")

    ws2.merge_cells('A1:H1')
    t2 = ws2['A1']
    t2.value     = "📋  Order Items Breakdown"
    t2.font      = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
    t2.fill      = fill(C_GOLD)
    t2.alignment = center()
    ws2.row_dimensions[1].height = 32

    headers2 = [
        ("Order ID",12),("Date",14),("Customer",18),
        ("Product",24),("Qty",7),("Unit Price",12),("Line Total",12),("Status",12)
    ]
    for col,(hdr,w) in enumerate(headers2,1):
        c = ws2.cell(row=2,column=col,value=hdr)
        c.font=hdr_font(); c.fill=fill(C_DARK); c.alignment=center(); c.border=border
        ws2.column_dimensions[get_column_letter(col)].width=w
    ws2.row_dimensions[2].height=22

    r = 3
    for order in qs:
        user   = order.user
        name   = f"{getattr(user,'first_name','')} {getattr(user,'last_name','')}".strip() or user.username
        try:   created = order.created_at.strftime('%d %b %Y')
        except: created = '—'
        try:   order_status = order.status
        except: order_status = '—'

        items = []
        try:   items = list(order.items.select_related('product').all())
        except:
            try: items = list(order.orderitem_set.select_related('product').all())
            except: pass

        if not items:
            # Write one row for order without item detail
            row_data = [f"#{order.id}", created, name, '(no items)', '—', '—', '—', order_status.capitalize()]
            bg = C_LIGHT if r % 2 == 0 else "FFFFFF"
            for col, val in enumerate(row_data, 1):
                c = ws2.cell(row=r, column=col, value=val)
                c.font=cell_font(); c.fill=fill(bg); c.border=border; c.alignment=center()
            r += 1
            continue

        for item in items:
            try:   pname = item.product.title
            except:
                try:   pname = item.product.name
                except: pname = '—'
            try:   qty   = item.quantity
            except: qty  = '—'
            try:   price = float(item.price)
            except:
                try:   price = float(item.product.price)
                except: price = 0
            try:   total = f"{qty * price:,.2f}" if isinstance(qty,int) else '—'
            except: total = '—'

            bg = C_LIGHT if r % 2 == 0 else "FFFFFF"
            row_data = [f"#{order.id}", created, name, pname, qty,
                        f"{price:,.2f}", total, order_status.capitalize()]
            for col, val in enumerate(row_data, 1):
                c = ws2.cell(row=r, column=col, value=val)
                c.font=cell_font(); c.fill=fill(bg); c.border=border; c.alignment=center()
            ws2.cell(row=r,column=8).fill = fill(status_color(order_status))
            ws2.cell(row=r,column=8).font = Font(name='Calibri',size=10,bold=True,color="FFFFFF")
            ws2.row_dimensions[r].height = 20
            r += 1

    ws2.freeze_panes = 'A3'

    # ══════════════════════════════════════════
    # SHEET 3 — Customers
    # ══════════════════════════════════════════
    from django.contrib.auth.models import User as DjangoUser
    ws3 = wb.create_sheet("Customers")

    ws3.merge_cells('A1:G1')
    t3 = ws3['A1']
    t3.value=f"👥  Registered Customers  |  {datetime.now().strftime('%d %b %Y')}"
    t3.font=Font(name='Calibri',size=13,bold=True,color="FFFFFF")
    t3.fill=fill(C_MID); t3.alignment=center()
    ws3.row_dimensions[1].height=30

    h3=[("ID",6),("Username",18),("First Name",16),("Last Name",16),
        ("Email",26),("Phone",14),("Joined",16)]
    for col,(hdr,w) in enumerate(h3,1):
        c=ws3.cell(row=2,column=col,value=hdr)
        c.font=hdr_font(); c.fill=fill(C_DARK); c.alignment=center(); c.border=border
        ws3.column_dimensions[get_column_letter(col)].width=w
    ws3.row_dimensions[2].height=22

    customers = DjangoUser.objects.filter(is_staff=False).select_related('profile').order_by('-date_joined')
    for ri,u in enumerate(customers,3):
        phone=''
        try:   phone=u.profile.phone
        except: pass
        bg = C_LIGHT if ri%2==0 else "FFFFFF"
        row_data=[u.id,u.username,u.first_name,u.last_name,u.email,phone,
                  u.date_joined.strftime('%d %b %Y')]
        for col,val in enumerate(row_data,1):
            c=ws3.cell(row=ri,column=col,value=val)
            c.font=cell_font(); c.fill=fill(bg); c.border=border; c.alignment=center()
        ws3.row_dimensions[ri].height=18

    ws3.freeze_panes='A3'

    # ── Save to response ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"P2P_Orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
