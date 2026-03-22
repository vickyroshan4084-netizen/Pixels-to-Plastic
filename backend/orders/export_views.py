"""
orders/export_views.py — ADD as new file in backend/orders/
─────────────────────────────────────────────────────────────
Then add to backend/orders/urls.py:

    from .export_views import export_orders_excel
    path('export/excel/', export_orders_excel, name='export-orders'),
"""
import io
from datetime import datetime, timedelta
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_orders_excel(request):
    if not request.user.is_staff:
        return Response({'error': 'Admin access required.'}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'error': 'Run: pip install openpyxl'}, status=500)

    try:
        from orders.models import Order
    except ImportError:
        return Response({'error': 'Order model not found.'}, status=500)

    # ── Build queryset ────────────────────────────────────────────────────────
    qs = Order.objects.all().select_related('user').order_by('-created_at')
    if request.GET.get('status'):
        qs = qs.filter(status=request.GET['status'])
    if request.GET.get('days'):
        try:
            qs = qs.filter(created_at__gte=datetime.now() - timedelta(days=int(request.GET['days'])))
        except ValueError:
            pass

    # ── Styles ────────────────────────────────────────────────────────────────
    def fill(c):  return PatternFill(fill_type='solid', fgColor=c)
    def hf(**k):  return Font(name='Calibri', size=k.get('s',11), bold=k.get('b',True),  color=k.get('c','FFFFFF'))
    def df(**k):  return Font(name='Calibri', size=k.get('s',10), bold=k.get('b',False), color=k.get('c','1A1A1A'))
    thin = Side(style='thin', color='DDDDDD')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    SCLR = {'completed':'27AE60','delivered':'27AE60','paid':'27AE60',
             'pending':'F39C12','processing':'2980B9','cancelled':'E74C3C','failed':'E74C3C'}

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ─ Sheet 1: Orders ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Orders'
    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = f'Pixels to Plastic — Orders Export  |  {datetime.now().strftime("%d %b %Y, %I:%M %p")}'
    c.font = hf(s=13); c.fill = fill('E8533C'); c.alignment = CTR
    ws.row_dimensions[1].height = 34

    hdrs = [('Order ID',10),('Date',16),('Customer',20),('Email',26),
            ('Phone',14),('Items',7),('Total ₹',12),('Payment',14),('Status',13)]
    for ci,(h,w) in enumerate(hdrs,1):
        c = ws.cell(row=2,column=ci,value=h)
        c.font=hf(); c.fill=fill('1A1A2E'); c.alignment=CTR; c.border=brd
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=22

    for ri, order in enumerate(qs, 3):
        u    = order.user
        name = f"{getattr(u,'first_name','')} {getattr(u,'last_name','')}".strip() or u.username
        phone = ''
        try:    phone = u.profile.phone
        except: pass
        try:    cnt = order.items.count()
        except: cnt = '?'
        try:    total = f"{float(order.total_price):,.2f}"
        except:
            try: total = f"{float(order.total):,.2f}"
            except: total = '—'
        try:    pay = order.payment_method
        except: pay = '—'
        try:    st = order.status
        except: st = '—'
        try:    dt = order.created_at.strftime('%d %b %Y  %H:%M')
        except: dt = '—'

        bg = 'F5F5F5' if ri % 2 == 0 else 'FFFFFF'
        for ci, val in enumerate([f'#{order.id}', dt, name, u.email, phone, cnt, total, pay, st.capitalize()], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font=df(); c.fill=fill(bg); c.border=brd
            c.alignment = CTR if ci in (1,5,6,7,8,9) else LFT
        # Status colour
        sc = ws.cell(row=ri, column=9)
        sc.fill = fill(SCLR.get((st or '').lower(), '7F8C8D'))
        sc.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.row_dimensions[ri].height = 19

    # Totals row
    last = qs.count() + 3
    ws.merge_cells(f'A{last}:I{last}')
    c = ws[f'A{last}']
    c.value = f'Total orders: {qs.count()}'
    c.font = hf(s=10); c.fill = fill('1A1A2E'); c.alignment = LFT
    ws.freeze_panes = 'A3'

    # ─ Sheet 2: Item Details ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Item Details')
    ws2.merge_cells('A1:H1')
    c = ws2['A1']
    c.value = 'Order Items Breakdown'
    c.font = hf(s=12); c.fill = fill('FFBE0B'); c.alignment = CTR
    ws2.row_dimensions[1].height = 30

    h2 = [('Order ID',10),('Date',14),('Customer',20),('Product',28),
          ('Qty',7),('Unit ₹',12),('Total ₹',12),('Status',13)]
    for ci,(h,w) in enumerate(h2,1):
        c = ws2.cell(row=2,column=ci,value=h)
        c.font=hf(); c.fill=fill('1A1A2E'); c.alignment=CTR; c.border=brd
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[2].height=22

    r = 3
    for order in qs:
        u    = order.user
        name = f"{getattr(u,'first_name','')} {getattr(u,'last_name','')}".strip() or u.username
        try:    dt = order.created_at.strftime('%d %b %Y')
        except: dt = '—'
        try:    st = order.status
        except: st = '—'
        items = []
        try:    items = list(order.items.select_related('product').all())
        except:
            try: items = list(order.orderitem_set.select_related('product').all())
            except: pass

        if not items:
            bg = 'F5F5F5' if r % 2 == 0 else 'FFFFFF'
            for ci, val in enumerate([f'#{order.id}', dt, name, '(no items)', '—', '—', '—', st.capitalize()], 1):
                c = ws2.cell(row=r,column=ci,value=val)
                c.font=df(); c.fill=fill(bg); c.border=brd; c.alignment=CTR
            r += 1; continue

        for item in items:
            try:    pname = item.product.title
            except:
                try: pname = item.product.name
                except: pname = '—'
            try:    qty = item.quantity
            except: qty = 1
            try:    price = float(item.price)
            except:
                try: price = float(item.product.price)
                except: price = 0.0
            line = f"{qty * price:,.2f}" if isinstance(qty, int) else '—'
            bg = 'F5F5F5' if r % 2 == 0 else 'FFFFFF'
            for ci, val in enumerate([f'#{order.id}', dt, name, pname, qty, f'{price:,.2f}', line, st.capitalize()], 1):
                c = ws2.cell(row=r,column=ci,value=val)
                c.font=df(); c.fill=fill(bg); c.border=brd; c.alignment=CTR
            ws2.cell(row=r,column=8).fill = fill(SCLR.get((st or '').lower(), '7F8C8D'))
            ws2.cell(row=r,column=8).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            ws2.row_dimensions[r].height = 19
            r += 1

    ws2.freeze_panes = 'A3'

    # ─ Sheet 3: Customers ─────────────────────────────────────────────────────
    from django.contrib.auth.models import User as DUser
    ws3 = wb.create_sheet('Customers')
    ws3.merge_cells('A1:H1')
    c = ws3['A1']
    c.value = f'Registered Customers  |  {datetime.now().strftime("%d %b %Y")}'
    c.font=hf(s=12); c.fill=fill('2C3E50'); c.alignment=CTR
    ws3.row_dimensions[1].height=28

    h3=[('ID',6),('Username',18),('Name',22),('Email',26),('Phone',14),('City',14),('Joined',16)]
    for ci,(h,w) in enumerate(h3,1):
        c=ws3.cell(row=2,column=ci,value=h)
        c.font=hf(); c.fill=fill('1A1A2E'); c.alignment=CTR; c.border=brd
        ws3.column_dimensions[get_column_letter(ci)].width=w
    ws3.row_dimensions[2].height=22

    for ri, u in enumerate(DUser.objects.filter(is_staff=False).select_related('profile').order_by('-date_joined'), 3):
        phone=''; city=''
        try:  phone=u.profile.phone; city=u.profile.city
        except: pass
        name = f"{u.first_name} {u.last_name}".strip() or u.username
        bg = 'F5F5F5' if ri%2==0 else 'FFFFFF'
        for ci, val in enumerate([u.id, u.username, name, u.email, phone, city, u.date_joined.strftime('%d %b %Y')], 1):
            c=ws3.cell(row=ri,column=ci,value=val)
            c.font=df(); c.fill=fill(bg); c.border=brd; c.alignment=CTR
        ws3.row_dimensions[ri].height=18
    ws3.freeze_panes='A3'

    # ── Send response ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname    = f"P2P_Orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
